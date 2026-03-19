from pathlib import Path
from openpyxl import Workbook, load_workbook

EXPORT_DIR = Path("exports")
EXPORT_FILE = EXPORT_DIR / "approved_invoices.xlsx"


def append_invoice_to_excel(payload: dict):
    print("APPEND TO EXCEL IS RUNNING")
    print("Export file path:", EXPORT_FILE.resolve())

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    headers = [
        "record_id",
        "vendor_name",
        "invoice_number",
        "invoice_date",
        "total",
        "currency",
        "status",
        "confidence_score",
    ]

    extraction = payload.get("extraction", {})
    workflow = payload.get("workflow", {})

    row = [
        payload["document"]["record_id"],
        extraction.get("vendor_name"),
        extraction.get("invoice_number"),
        extraction.get("invoice_date"),
        extraction.get("total"),
        extraction.get("currency"),
        workflow.get("status"),
        workflow.get("confidence_score"),
    ]

    try:
        if EXPORT_FILE.exists():
            workbook = load_workbook(EXPORT_FILE)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Approved Invoices"
            worksheet.append(headers)

        worksheet.append(row)
        workbook.save(EXPORT_FILE)
        print("EXCEL FILE SAVED SUCCESSFULLY")

    except PermissionError:
        raise PermissionError(
            f"Cannot write to export file: {EXPORT_FILE}. Please close the Excel file and try again."
        )